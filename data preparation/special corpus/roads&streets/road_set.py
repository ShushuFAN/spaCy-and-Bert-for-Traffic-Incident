import pandas as pd
import openpyxl


def data_set(path, sheet_name, col, list_data):
    print("parsing %s"%path)
    data_col=openpyxl.load_workbook(path)[sheet_name]
    for i in range(2,data_col.max_row+1):
        # print(Street.cell(i,col).value)
        data=data_col.cell(i, col).value.lower()
        if data not in ["", '–９９','-99','－９９']:
            list_data.append(data)
    return list_data


if __name__ == '__main__':
    data_path1=["PSI_Street Name_201712.xlsx","工作表1",1]
    data_path2=["PSI_Street Name_201806.xlsx","Main",1]
    data_path3=["PSI_Street Name_201812.xlsx","Main",1]
    data_path4=["PSI_Street Name_201906.xlsx","PSI",1]
    data_path5=["PSI_Street Name_201912.xlsx","PSI",1]
    data_path6=["PSI_Street Name_202006.xlsx","PSI",1]
    data_path7=["road_centerline.xlsx","Sheet1",2]
    data_list=[data_path1,data_path2,data_path3,data_path4,data_path5,data_path6,data_path7]
    list_data = []
    for data in data_list:
        list_data=data_set(data[0], data[1], data[2], list_data)
    supplement=["HEUNG YUEN WAI HIGHWAY","CHING LAI ROAD","HUNG HOM BYPASS","LUNG HOP STREET","TUNG LEI PATH","HUNG LENG NORTH ROAD","MUSEUM DRIVE","CHEUNG SHAN TUNNEL","LUNG SHAN TUNNEL"]
    for i in supplement:
        list_data.append(i.lower())
    temp = list(set(list_data))
    print("writing text")
    road_dic = open('road_dic.txt', 'a', encoding='utf-8')
    for line in temp:
        road_dic.write(line + "\n")
    road_dic.close()